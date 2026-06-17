import argparse
import csv
import json
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl.styles import Font

ROOT_DIR = Path(__file__).resolve().parent
DEFAULT_GT_FILE = 'ground_truth.jsonl'
DEFAULT_OUTPUT_DIR = 'output'
DEFAULT_CSV = 'output/eval_summary.csv'
DEFAULT_XLSX = 'output/eval_summary.xlsx'
NUM_DIMS = 23


def _resolve_path(path: str) -> Path:
    """relative to root directory"""
    p = Path(path)
    return p if p.is_absolute() else ROOT_DIR / p

DIMENSION_GROUPS = [
    ("Medical Interviewing", list(range(0, 8))),
    ("Humanistic Care", list(range(8, 16))),
    ("Diagnosis and Treatment Management", list(range(16, 23))),
]

FAILED_PRED = object()


def load_ground_truth(path: Path) -> Dict[int, List[int]]:
    gt: Dict[int, List[int]] = {}
    with path.open(encoding="utf-8") as f:
        for line in f:
            row = json.loads(line)
            rid = row.get("id")
            out = row.get("output")
            if rid is None or not isinstance(out, list) or len(out) != NUM_DIMS:
                continue
            gt[rid] = [int(x) for x in out]
    return gt


def _normalize_pred(output) -> Optional[List[int]]:
    if not isinstance(output, list) or len(output) != NUM_DIMS:
        return None
    try:
        return [int(x) for x in output]
    except (TypeError, ValueError):
        return None


def load_predictions(path: Path) -> Tuple[Dict[int, object], dict]:
    """return (id -> pred or FAILED_PRED), statistics"""
    preds: Dict[int, object] = {}
    stats = {
        "n_lines": 0,
        "n_valid": 0,
        "n_failed": 0,
        "n_error": 0,
        "n_parse_error": 0,
        "n_bad_output": 0,
    }
    with path.open(encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            stats["n_lines"] += 1
            row = json.loads(line)
            rid = row.get("id")
            if rid is None:
                continue
            if "error" in row:
                stats["n_error"] += 1
                stats["n_failed"] += 1
                preds[rid] = FAILED_PRED
                continue
            if "parse_error" in row:
                stats["n_parse_error"] += 1
                stats["n_failed"] += 1
                preds[rid] = FAILED_PRED
                continue
            pred = _normalize_pred(row.get("output"))
            if pred is None:
                stats["n_bad_output"] += 1
                stats["n_failed"] += 1
                preds[rid] = FAILED_PRED
                continue
            preds[rid] = pred
            stats["n_valid"] += 1
    return preds, stats


def binary_f1(tp: int, fp: int, fn: int) -> float:
    if tp == 0 and fp == 0 and fn == 0:
        return 1.0
    precision = tp / (tp + fp) if (tp + fp) > 0 else 0.0
    recall = tp / (tp + fn) if (tp + fn) > 0 else 0.0
    if precision + recall == 0:
        return 0.0
    return 2 * precision * recall / (precision + recall)


def _dim_confusion(
    gt: Dict[int, List[int]],
    preds: Dict[int, object],
    common_ids: List[int],
    dim: int,
) -> Tuple[int, int, int, int, int]:
    """return (tp, fp, fn, correct, total) single dimension statistics"""
    tp = fp = fn = correct = 0
    for rid in common_ids:
        y_true = gt[rid][dim]
        pred = preds[rid]
        if pred is FAILED_PRED:
            if y_true == 1:
                fn += 1
            else:
                fp += 1
        else:
            y_pred = pred[dim]
            if y_true == 1 and y_pred == 1:
                tp += 1
            elif y_true == 0 and y_pred == 1:
                fp += 1
            elif y_true == 1 and y_pred == 0:
                fn += 1
            if y_true == y_pred:
                correct += 1
    return tp, fp, fn, correct, len(common_ids)


def compute_metrics(
    gt: Dict[int, List[int]], preds: Dict[int, object]
) -> Optional[dict]:
    common_ids = sorted(set(gt.keys()) & set(preds.keys()))
    if not common_ids:
        return None

    dim_f1: List[float] = []
    dim_correct: List[int] = []
    dim_total: List[int] = []

    for dim in range(NUM_DIMS):
        tp, fp, fn, correct, total = _dim_confusion(gt, preds, common_ids, dim)
        dim_f1.append(binary_f1(tp, fp, fn))
        dim_correct.append(correct)
        dim_total.append(total)

    macro_f1 = sum(dim_f1) / NUM_DIMS
    hamming_accuracy = (
        sum(dim_correct) / sum(dim_total) if sum(dim_total) else 0.0
    )

    groups = {}
    for name, dims in DIMENSION_GROUPS:
        group_f1 = sum(dim_f1[d] for d in dims) / len(dims)
        group_correct = sum(dim_correct[d] for d in dims)
        group_total = sum(dim_total[d] for d in dims)
        groups[name] = {
            "macro_f1": group_f1,
            "hamming_accuracy": group_correct / group_total if group_total else 0.0,
        }

    return {
        "macro_f1": macro_f1,
        "hamming_accuracy": hamming_accuracy,
        "groups": groups,
        "n_evaluated": len(common_ids),
        "n_gt": len(gt),
    }


def discover_model_files(output_dir: Path, models: Optional[List[str]] = None) -> List[Path]:
    files = []
    for p in sorted(output_dir.glob("*.jsonl")):
        if models and p.stem not in models:
            continue
        files.append(p)
    return files


def evaluate_all(
    gt_path: Path,
    output_dir: Path,
    models: Optional[List[str]] = None,
) -> Tuple[List[dict], List[dict]]:
    gt = load_ground_truth(gt_path)
    summary_rows: List[dict] = []
    category_rows: List[dict] = []
    for pred_path in discover_model_files(output_dir, models):
        preds, stats = load_predictions(pred_path)
        metrics = compute_metrics(gt, preds)
        row = {
            "model": pred_path.stem,
            "macro_f1": "",
            "hamming_accuracy": "",
            "n_evaluated": 0,
            "n_gt": len(gt),
            "n_lines": stats["n_lines"],
            "n_valid_pred": stats["n_valid"],
            "n_failed": stats["n_failed"],
            "n_error": stats["n_error"],
            "n_parse_error": stats["n_parse_error"],
            "n_bad_output": stats["n_bad_output"],
        }
        cat_row = {"model": pred_path.stem, "n_valid_pred": stats["n_valid"]}
        if metrics:
            row["macro_f1"] = round(metrics["macro_f1"], 6)
            row["hamming_accuracy"] = round(metrics["hamming_accuracy"], 6)
            row["n_evaluated"] = metrics["n_evaluated"]
            cat_row["Overall Macro-F1"] = round(metrics["macro_f1"], 6)
            cat_row["Overall Hamming Acc"] = round(metrics["hamming_accuracy"], 6)
            for name, _ in DIMENSION_GROUPS:
                g = metrics["groups"][name]
                cat_row[f"{name} Macro-F1"] = round(g["macro_f1"], 6)
                cat_row[f"{name} Hamming Acc"] = round(g["hamming_accuracy"], 6)
        else:
            cat_row["Overall Macro-F1"] = ""
            cat_row["Overall Hamming Acc"] = ""
            for name, _ in DIMENSION_GROUPS:
                cat_row[f"{name} Macro-F1"] = ""
                cat_row[f"{name} Hamming Acc"] = ""
        summary_rows.append(row)
        category_rows.append(cat_row)

    sort_key = lambda r: (-float(r.get("Overall Macro-F1") or r.get("macro_f1") or 0), r["model"])
    summary_rows.sort(key=lambda r: (-float(r["macro_f1"] or 0), r["model"]))
    category_rows.sort(key=sort_key)
    return summary_rows, category_rows


def _format_category_sheet(ws, cat_columns: List[str]) -> None:
    bold_font = Font(bold=True)
    underline_font = Font(underline="single")

    for col_idx, col_name in enumerate(cat_columns, start=1):
        if col_name == "model":
            continue

        value_rows: Dict[float, List[int]] = {}
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            val = cell.value
            if val is None or val == "":
                continue
            try:
                fval = float(val)
            except (TypeError, ValueError):
                continue
            value_rows.setdefault(fval, []).append(row)

        if not value_rows:
            continue

        ranked = sorted(value_rows.keys(), reverse=True)
        for row in value_rows[ranked[0]]:
            ws.cell(row=row, column=col_idx).font = bold_font
        if len(ranked) >= 2:
            for row in value_rows[ranked[1]]:
                ws.cell(row=row, column=col_idx).font = underline_font


def write_excel(summary_rows: List[dict], category_rows: List[dict], xlsx_path: Path) -> None:
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    cat_columns = ["model"]
    for name, _ in DIMENSION_GROUPS:
        cat_columns.extend([f"{name} Macro-F1", f"{name} Hamming Acc"])
    cat_columns.extend(["Overall Macro-F1", "Overall Hamming Acc", "n_valid_pred"])

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Summary", index=False)
        pd.DataFrame(category_rows, columns=cat_columns).to_excel(
            writer, sheet_name="By Category", index=False
        )
        _format_category_sheet(writer.sheets["By Category"], cat_columns)


def write_csv(rows: List[dict], csv_path: Path) -> None:
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "model",
        "macro_f1",
        "hamming_accuracy",
        "n_evaluated",
        "n_gt",
        "n_lines",
        "n_valid_pred",
        "n_failed",
        "n_error",
        "n_parse_error",
        "n_bad_output",
    ]
    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def print_table(rows: List[dict]) -> None:
    print(f"{'model':<40} {'macro_f1':>10} {'hamming_acc':>12} {'n_eval':>8}")
    print("-" * 72)
    for r in rows:
        mf = r["macro_f1"] if r["macro_f1"] != "" else "N/A"
        ha = r["hamming_accuracy"] if r["hamming_accuracy"] != "" else "N/A"
        print(f"{r['model']:<40} {str(mf):>10} {str(ha):>12} {r['n_evaluated']:>8}")


def parse_args():
    parser = argparse.ArgumentParser(description="对照 ground truth 计算 Macro-F1 与 Hamming Accuracy")
    parser.add_argument(
        "--gt",
        default=DEFAULT_GT_FILE,
        help=f"ground truth jsonl (default: {DEFAULT_GT_FILE}, relative to project directory)",
    )
    parser.add_argument(
        "--output-dir",
        default=DEFAULT_OUTPUT_DIR,
        help=f"model output directory (default: {DEFAULT_OUTPUT_DIR}, relative to project directory)",
    )
    parser.add_argument(
        "-o", "--csv",
        default=DEFAULT_CSV,
        help=f"output CSV path (default: {DEFAULT_CSV}, relative to project directory)",
    )
    parser.add_argument(
        "--xlsx",
        default=DEFAULT_XLSX,
        help=f"output Excel path (default: {DEFAULT_XLSX}, relative to project directory)",
    )
    parser.add_argument(
        "-m", "--models",
        nargs="+",
        metavar="MODEL",
        help="evaluate only specified models (file name does not contain .jsonl)",
    )
    return parser.parse_args()


def main():
    args = parse_args()
    gt_path = _resolve_path(args.gt)
    output_dir = _resolve_path(args.output_dir)
    csv_path = _resolve_path(args.csv)
    xlsx_path = _resolve_path(args.xlsx)

    if not gt_path.is_file():
        raise SystemExit(f"ground truth file does not exist: {gt_path}")

    summary_rows, category_rows = evaluate_all(gt_path, output_dir, args.models)
    if not summary_rows:
        raise SystemExit(f"no evaluatable model jsonl found in {output_dir}")

    write_csv(summary_rows, csv_path)
    write_excel(summary_rows, category_rows, xlsx_path)
    print(f"GT: {gt_path} ({summary_rows[0]['n_gt']} records)")
    print(f"written to CSV: {csv_path}")
    print(f"written to Excel: {xlsx_path} (Summary + By Category)\n")
    print_table(summary_rows)


if __name__ == "__main__":
    main()
